import openpyxl
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

#Criando a planilha que vai receber os dados (wb)
wb = openpyxl.Workbook()
wb.create_sheet('Ferramentas Cadastradas')
sheet = wb['Ferramentas Cadastradas']

def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 30

    sheet.cell(row=1, column=1).value = "Descrição da ferramenta"
    sheet.cell(row=1, column=2).value = "Fabricante"
    sheet.cell(row=1, column=3).value = "Voltagem "
    sheet.cell(row=1, column=4).value = "Part Number"
    sheet.cell(row=1, column=5).value = "Tamanho"
    sheet.cell(row=1, column=6).value = "Tipo de Ferramenta"
    sheet.cell(row=1, column=7).value = "Material da Ferramenta"
    sheet.cell(row=1, column=8).value = "Máximo de Reserva (em horas)"

def focus1(evento):
    descricao.focus_set()


def focus2(evento):
    fabricante.focus_set()


def focus3(evento):
    voltagem.focus_set()


def focus4(evento):
    part_number.focus_set()


def focus5(evento):
    tamanho.focus_set()


def focus6(evento):
    tipo.focus_set()

def focus7(evento):
    material.focus_set()

def focus8(evento):
    maximo_reserva.focus_set()


def clear():
    descricao.delete(0, END)
    fabricante.delete(0, END)
    voltagem.delete(0, END)
    part_number.delete(0, END)
    tamanho.delete(0, END)
    tipo.delete(0, END)
    material.delete(0, END)
    maximo_reserva.delete(0, END)


def inserir():
    if (descricao.get() == "" or descricao.get() == " " and
            fabricante.get() == "" or fabricante.get() == " " and
            voltagem.get() == "" or voltagem.get() == " " and
            part_number.get() == "" or part_number.get() == " " and
            tamanho.get() == "" or tamanho.get() == " " and
            tipo.get() == "" or tipo.get() == " " and
            material.get() == "" or material.get() == " " and
            maximo_reserva.get() == "" or maximo_reserva.get() == " "):

        messagebox.showerror("Error", "Voce deve preencher todos os campos")
    elif (voltagem.get() != '110V' and voltagem.get() != "220V" and voltagem.get() != "360V" and voltagem.get() != "110~220V"):
        messagebox.showerror("Error", "Essa voltagem não existe, favor preencher corretamente")
    elif (tamanho.get() != 'Polegadas' and tamanho.get() != "MM" and tamanho.get() != "CM" and tamanho.get() != "M"):
        messagebox.showerror("Error", "Tamanho não aceitável, favor preencher corretamente")
    else:
        resposta = messagebox.askquestion("Tem certeza?", "Cadastrar ferramenta " + descricao.get() + "?")
        if resposta == "yes":
            messagebox.showinfo("Sucesso", "Ferramenta cadastrada com sucesso!")
            #wb = openpyxl.load_workbook('testetecnico1.xlsx')
            sheet = wb['Ferramentas Cadastradas']
            linha_atual = sheet.max_row
            coluna_atual = sheet.max_column

            sheet.cell(row=linha_atual + 1, column=1).value = descricao.get()
            sheet.cell(row=linha_atual + 1, column=2).value = fabricante.get()
            sheet.cell(row=linha_atual + 1, column=3).value = voltagem.get()
            sheet.cell(row=linha_atual + 1, column=4).value = part_number.get()
            sheet.cell(row=linha_atual + 1, column=5).value = unidadeMedida.get() + " " + tamanho.get()
            sheet.cell(row=linha_atual + 1, column=6).value = tipo.get()
            sheet.cell(row=linha_atual + 1, column=7).value = material.get()
            sheet.cell(row=linha_atual + 1, column=8).value = maximo_reserva.get()
            wb.save('testeferramentas1.xlsx')
            descricao.focus_set()
        else:
            messagebox.showwarning("Cancelado", "Ferramenta não cadastrado!")
        clear()


if __name__ == "__main__":
    root = Tk()

    root.configure(background='light blue')

    root.title("Cadastro de Ferramentas")
    root.geometry("900x300")

    excel()

    titulo = Label(root, text="Cadastro de Ferramentas", bg="light blue")

    descricao = Label(root, text="Descrição da Ferramenta", bg="light blue")

    fabricante = Label(root, text="Fabricante", bg="light blue")

    voltagem = Label(root, text="Voltagem", bg="light blue")

    part_number = Label(root, text="Part Number", bg="light blue")

    tamanho = Label(root, text="Tamanho", bg="light blue")

    unidadeMedida = Label(root, text="Unidade de medida", bg="light blue")

    tipo = Label(root, text="Tipo de Ferramenta", bg="light blue")

    material = Label(root, text="Material", bg="light blue")

    maximo_reserva = Label(root, text="Máximo de Reserva (em horas)", bg="light blue")

    titulo.grid(row=0, column=1)
    descricao.grid(row=1, column=0)
    fabricante.grid(row=2, column=0)
    voltagem.grid(row=3, column=0)
    part_number.grid(row=4, column=0)
    tamanho.grid(row=5, column=0)
    unidadeMedida.grid(row=5, column=3)
    tipo.grid(row=6, column=0)
    material.grid(row=7, column=0)
    maximo_reserva.grid(row=8, column=0)

    descricao = Entry(root)
    fabricante = Entry(root)
    voltagem = Entry(root)
    part_number = Entry(root)
    tamanho = Entry(root)
    tipo = Entry(root)
    material = Entry(root)
    maximo_reserva = Entry(root)
    unidadeMedida = Entry(root)

    descricao.bind("<Return>", focus1)

    fabricante.bind("<Return>", focus2)

    voltagem.bind("<Return>", focus3)

    part_number.bind("<Return>", focus4)

    tamanho.bind("<Return>", focus5)

    tipo.bind("<Return>", focus6)

    material.bind("<Return>", focus6)

    maximo_reserva.bind("<Return>", focus6)

    descricao.grid(row=1, column=1, ipadx="100")
    fabricante.grid(row=2, column=1, ipadx="100")
    voltagemEscolha = ["110V", "220V", "360V", "110V~220V"]
    voltagem = ttk.Combobox(root, values=voltagemEscolha)
    voltagem.set("110V")
    voltagem.grid(row=3, column=1, ipadx="90")
    part_number.grid(row=4,column=1,ipadx="100")
    tamanhoEscolha = ["Polegadas", "MM", "CM", "M"]
    tamanho = ttk.Combobox(root, values=tamanhoEscolha)
    tamanho.set("Polegadas")
    tamanho.grid(row=5, column=4, ipadx="0")
    unidadeMedida.grid(row=5, column=1,ipadx="100")
    tipo.grid(row=6, column=1, ipadx="100")
    material.grid(row=7, column=1, ipadx="100")
    maximo_reserva.grid(row=8, column=1, ipadx="100")

    excel()

    cadastrar = Button(root, text="Cadastrar", fg="White",
                    bg="Black", command=inserir)
    cadastrar.grid(row=10, column=1)

    root.mainloop()